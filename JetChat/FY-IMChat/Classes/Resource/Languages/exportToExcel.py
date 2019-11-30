# -*- coding: utf-8 -*-
import codecs
import xlwt
import os
import sys
from openpyxl import Workbook
#from imp import reload
#
#reload(sys)

##########################################################
def readKeysFromFilePath(path):
    listKey = []
    for string in codecs.open(path,'r','utf-8').readlines():
        list = string.split('=')
        if len(list) >= 2:
            # k = list[0].strip().lstrip('"').rstrip('"')
            a = list[0]
            if a.startswith('//'):
                continue
            k = a[a.find('"')+1: a.rfind('"')]
            if len(k) == 0:
                continue
            listKey.append(k)
    return listKey
##########################################################
def readValuesFromFilePath(path):
    listValue = []
    for string in codecs.open(path,'r','utf-8').readlines():
    	list = string.split('=')
    	if len(list) >= 2:
            # v = list[1].strip().lstrip('"').rstrip('\n').rstrip().rstrip(';').rstrip().rstrip('"')
            b = list[1]
            v = b[b.find('"')+1: b.rfind('"')]
            listValue.append(v)

    return listValue
##########################################################
def readKVFromFilePath(path):
    kv = dict()
    for string in codecs.open(path,'r','utf-8').readlines():
        list = string.split('=')
        if len(list) >= 2:
            a = list[0]
            if a.startswith('//'):
                continue
            k = a[a.find('"')+1: a.rfind('"')]
            if len(k) == 0:
                continue
            b = list[1]
            v = b[b.find('"')+1: b.rfind('"')]
            kv[k] = v
    return kv
##########################################################

def main():
    fileDict = dict()
    for file in os.listdir('./'):
        if file.__contains__('.lproj'):
            if os.path.isdir(file):
                for i in os.listdir(file):
                    if i.__contains__('Localizable'):
                        fileDict[file] = i
                        break


    
    fileName = 'LauguageExcel.xls'
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Localizable")
#    wb = Workbook()
#    ws = wb.active

    # tall_style = xlwt.easyxf('font:height 720;')
    width = 256*20
    
#    ws.cell(row=1, column=1).value = 'key'
    ws.write(0, 0, 'key')
    ws.col(0).width = width
    path = 'zh-Hans.lproj/Localizable.strings'
    keys = readKeysFromFilePath(path)
    for x in range(len(keys)):
#        ws.cell(row=x+2, column=1).value = keys[x]
        ws.write(x+1, 0, keys[x])



    for y in range(len(fileDict)):
        key = list(fileDict.keys())[y]
        loc = fileDict[key]
        path = key + "/" + loc
#        print(path)
#        ws.cell(row=1, column=y+2).value = key
        ws.write(0, y+1, key)
        ws.col(y+1).width = width
        kv = readKVFromFilePath(path)
        for x in range(len(keys)):
            key = keys[x]
            value = kv.get(key, '')
#            if len(value) == 0:
#                print(key)
#            ws.cell(row=x+2, column=y+2).value = value
            ws.write(x+1, y+1, value)


    wb.save(fileName)

if __name__ == '__main__':
    main()
