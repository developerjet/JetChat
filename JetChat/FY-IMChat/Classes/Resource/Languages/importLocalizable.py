# -*- coding: utf-8 -*-
import sys
import xdrlib
import xlrd
import os
import shutil
from openpyxl import load_workbook
#from imp import reload
#
#reload(sys)

def main():
    
    fileName = 'LauguageExcel.xls'
    #'localizableExcel.xls'
#    workbook = load_workbook(fileName)
#    sheet = workbook.worksheets[0]
    workbook = xlrd.open_workbook(fileName)
    sheet = workbook.sheet_by_index(0)
    
    fp = 0

    for y in range(sheet.ncols): ##sheet.max_column
        if y == 0:
            continue
        for x in range(sheet.nrows): ##sheet.max_row
            if x == 0:
                if fp != 0:
                    fp.close
                #file = sheet.cell(row=1, column=y).value + '/Localizable.strings'
                if sheet.cell_value(0, y) == 'zh-Hans.lproj':
                    break
#                if sheet.cell_value(0, y) == 'en.lproj':
#                    break
#                if sheet.cell_value(0, y) == 'Base.lproj':
#                    break
                file = sheet.cell_value(0, y) + '/Localizable.strings'
                fp = open(file, 'wb')
#                print(file)
            else:
#                key = sheet.cell(row=x, column=1).value
#                value = sheet.cell(row=x, column=y).value
#                if value == None:
#                    value = ""
#                kv = '"' + key + '" = "' + value + '";\n'
                kv = '"' + str(sheet.cell_value(x, 0)) + '" = "' + str(sheet.cell_value(x, y)) + '";\n'
                kv = kv.encode('utf-8')
                if x % 10 == 0:
                    fp.write('\n'.encode('utf-8'))
                fp.write(kv) 
    
    if fp != 0:
        fp.close
             

if __name__=="__main__":
    main()

    # table = data.sheets()[0]
    # colnames = table.row_values(0) #第一行数据
    
    # colKeys = table.col_values(0) #第一列key数据
    # colValues_zh_CN = table.col_values(1) #简体中文数据
    # colValues_English = table.col_values(2)#英文数据
    # nrows = len(colKeys) #总行数
    # ncols = len(colnames) #总列数
    
    # languageList = []
    # for indexCol in range(1,ncols):
    #     list = []
    #     colValues = table.col_values(indexCol)
    #     for indexRow in range(1,nrows):
            
    #         value = colValues[indexRow]
    #         value
    #         if (len(value)==0):
    #             value = colValues_English[indexRow]
    #         keyValue = '"' + colKeys[indexRow] + '"' + ' = ' + '"' + value + '"' + ';\n'
    #         list.append(keyValue)
    #     languageList.append(''.join(list))


    # for index in range(len(languageList)):
    #     print languageList[index]
    #     fileName = str(index) + 'Localizable.strings'
    #     os.system(r'touch %s' % fileName)
        
    #     fp = open(fileName,'wb+')
    #     fp.write(languageList[index])
    #     fp.close()

